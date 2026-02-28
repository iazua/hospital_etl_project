# =============================================
# LOAD.PY - Capa de Carga (L)
# Proyecto: Pipeline de Estadísticas Hospitalarias
# =============================================

import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from datetime import datetime

from config import PROCESSED_DIR, REPORTS_DIR


# ---------------------------------------------
# HELPERS
# ---------------------------------------------
def log(mensaje: str):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{timestamp}] {mensaje}")

def leer_processed(nombre: str) -> pd.DataFrame:
    ruta = os.path.join(PROCESSED_DIR, f"{nombre}.csv")
    return pd.read_csv(ruta, encoding="utf-8-sig")


# ---------------------------------------------
# ESTILOS EXCEL
# ---------------------------------------------
COLOR_HEADER        = "1F4E79"   # Azul oscuro institucional
COLOR_HEADER_FONT   = "FFFFFF"   # Blanco
COLOR_FILA_PAR      = "EBF3FB"   # Azul muy claro para filas alternas
COLOR_TOTAL         = "D6E4F0"   # Azul claro para filas de totales

BORDER_THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

def estilo_header(cell):
    cell.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
    cell.font      = Font(bold=True, color=COLOR_HEADER_FONT, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = BORDER_THIN

def estilo_celda(cell, fila_par: bool = False):
    if fila_par:
        cell.fill = PatternFill("solid", fgColor=COLOR_FILA_PAR)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border    = BORDER_THIN

def ajustar_columnas(ws, min_width=12, max_width=45):
    """Ajusta el ancho de cada columna según su contenido."""
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in col
        )
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, min_width), max_width)


# ---------------------------------------------
# FUNCIÓN PRINCIPAL DE ESCRITURA DE HOJAS
# ---------------------------------------------
def escribir_hoja(ws, df: pd.DataFrame, titulo: str):
    """
    Escribe un DataFrame en una hoja de Excel con:
    - Fila de título fusionada
    - Encabezados estilizados
    - Filas alternas coloreadas
    - Columnas auto-ajustadas
    """
    n_cols = len(df.columns)

    # -- Fila 1: Título de la hoja --
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    titulo_cell = ws.cell(row=1, column=1, value=titulo)
    titulo_cell.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
    titulo_cell.font      = Font(bold=True, color=COLOR_HEADER_FONT, size=13)
    titulo_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # -- Fila 2: Encabezados de columnas --
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name.replace("_", " ").title())
        estilo_header(cell)
    ws.row_dimensions[2].height = 25

    # -- Filas de datos --
    for row_idx, row in enumerate(df.itertuples(index=False), start=3):
        fila_par = (row_idx % 2 == 0)
        for col_idx, value in enumerate(row, start=1):
            # Convertir NaN a cadena vacía para Excel
            valor = "" if (isinstance(value, float) and pd.isna(value)) else value
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            estilo_celda(cell, fila_par)

    # -- Fila de totales (solo para columnas numéricas) --
    fila_total = len(df) + 3
    ws.cell(row=fila_total, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=fila_total, column=1).fill = PatternFill("solid", fgColor=COLOR_TOTAL)

    for col_idx, col_name in enumerate(df.columns, start=1):
        if pd.api.types.is_numeric_dtype(df[col_name]):
            cell = ws.cell(
                row=fila_total,
                column=col_idx,
                value=f"=SUM({get_column_letter(col_idx)}3:{get_column_letter(col_idx)}{fila_total - 1})"
            )
            cell.font  = Font(bold=True)
            cell.fill  = PatternFill("solid", fgColor=COLOR_TOTAL)
            cell.border = BORDER_THIN

    # Congelar paneles: encabezado siempre visible al hacer scroll
    ws.freeze_panes = "A3"

    ajustar_columnas(ws)


# ---------------------------------------------
# HOJA DE RESUMEN EJECUTIVO
# ---------------------------------------------
def escribir_resumen(ws, indicadores: dict):
    """Hoja inicial con métricas clave de un vistazo."""
    ws.title = "Resumen Ejecutivo"

    # Calcular KPIs principales
    df_ocup    = indicadores["ind_ocupacion_camas"]
    df_estadia = indicadores["ind_estadia_servicio"]
    df_egresos = indicadores["ind_egresos_mensuales"]
    df_atenc   = indicadores["ind_atenciones_mes"]

    tasa_ocup_promedio  = df_ocup["tasa_ocupacion_pct"].mean().round(1)
    promedio_estadia    = df_estadia["promedio_dias"].mean().round(1)
    total_egresos       = df_egresos["total_egresos"].sum()
    total_atenciones    = df_atenc["total_atenciones"].sum()
    fecha_reporte       = datetime.now().strftime("%d/%m/%Y %H:%M")

    kpis = [
        ("📅 Fecha de generación",          fecha_reporte),
        ("🏥 Tasa de ocupación promedio",    f"{tasa_ocup_promedio}%"),
        ("📆 Promedio días de estadía",      f"{promedio_estadia} días"),
        ("🚪 Total egresos (período)",       f"{total_egresos:,}"),
        ("🩺 Total atenciones ambulatorias", f"{total_atenciones:,}"),
    ]

    # Título
    ws.merge_cells("A1:B1")
    cell = ws["A1"]
    cell.value     = "RESUMEN EJECUTIVO — ESTADÍSTICAS HOSPITALARIAS"
    cell.font      = Font(bold=True, color=COLOR_HEADER_FONT, size=14)
    cell.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # KPIs
    for i, (indicador, valor) in enumerate(kpis, start=2):
        ws.cell(row=i, column=1, value=indicador).font = Font(bold=True, size=12)
        ws.cell(row=i, column=2, value=valor).font     = Font(size=12)
        ws.row_dimensions[i].height = 22

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 30


# ---------------------------------------------
# ORQUESTADOR DE CARGA
# ---------------------------------------------
def ejecutar_carga() -> str:
    log("=" * 50)
    log("Iniciando carga de datos (Capa L)")
    log("=" * 50)

    os.makedirs(REPORTS_DIR, exist_ok=True)
    fecha_str   = datetime.now().strftime("%Y%m%d")
    nombre_xlsx = f"reporte_estadisticas_{fecha_str}.xlsx"
    ruta_xlsx   = os.path.join(REPORTS_DIR, nombre_xlsx)

    # Leer todos los indicadores procesados
    indicadores = {
        "ind_ocupacion_camas":          leer_processed("ind_ocupacion_camas"),
        "ind_estadia_servicio":         leer_processed("ind_estadia_servicio"),
        "ind_estadia_diagnostico":      leer_processed("ind_estadia_diagnostico"),
        "ind_egresos_mensuales":        leer_processed("ind_egresos_mensuales"),
        "ind_atenciones_mes":           leer_processed("ind_atenciones_mes"),
        "ind_atenciones_especialidad":  leer_processed("ind_atenciones_especialidad"),
        "ind_distribucion_pacientes":   leer_processed("ind_distribucion_pacientes"),
    }

    # Definir nombre de hoja y título por indicador
    hojas = [
        ("ind_ocupacion_camas",         "Ocupación Camas",          "Tasa de Ocupación de Camas por Servicio"),
        ("ind_estadia_servicio",        "Estadía x Servicio",       "Promedio de Días de Estadía por Servicio"),
        ("ind_estadia_diagnostico",     "Estadía x Diagnóstico",    "Días de Estadía por Categoría Diagnóstica"),
        ("ind_egresos_mensuales",       "Egresos Mensuales",        "Egresos Hospitalarios por Mes y Servicio"),
        ("ind_atenciones_mes",          "Atenciones x Mes",         "Atenciones Ambulatorias por Mes y Tipo"),
        ("ind_atenciones_especialidad", "Atenciones x Especialidad","Atenciones Ambulatorias por Especialidad"),
        ("ind_distribucion_pacientes",  "Distribución Pacientes",   "Distribución por Previsión, Sexo y Grupo Etario"),
    ]

    # Crear Excel con openpyxl
    with pd.ExcelWriter(ruta_xlsx, engine="openpyxl") as writer:
        # Escribir hojas de indicadores
        for key, nombre_hoja, titulo in hojas:
            df = indicadores[key]
            df.to_excel(writer, sheet_name=nombre_hoja, index=False, startrow=1)
            ws = writer.sheets[nombre_hoja]
            escribir_hoja(ws, df, titulo)
            log(f"  ✓ Hoja '{nombre_hoja}' escrita → {len(df):,} registros.")

        # Escribir hoja de resumen al final y moverla al inicio
        wb = writer.book
        wb.create_sheet("Resumen Ejecutivo")
        ws_resumen = wb["Resumen Ejecutivo"]
        escribir_resumen(ws_resumen, indicadores)
        wb.move_sheet("Resumen Ejecutivo", offset=-len(hojas))
        log(f"  ✓ Hoja 'Resumen Ejecutivo' escrita.")

    log("=" * 50)
    log(f"Carga completada → {nombre_xlsx}")
    log(f"Ruta: {ruta_xlsx}")
    log("=" * 50)

    return ruta_xlsx


# ---------------------------------------------
# MAIN
# ---------------------------------------------
def main():
    ejecutar_carga()


if __name__ == "__main__":
    main()